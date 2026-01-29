"""
QLNNN Offline - Export Data Module
Export search results to Excel (XLSX)
"""

from typing import List, Dict, Any
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
import sys

sys.path.append(str(Path(__file__).parent.parent))

from utils.date_utils import format_date_vn
from config import STATUS_COLORS


# Column definitions for export
EXPORT_COLUMNS = [
    ("ho_ten", "Họ và tên"),
    ("ngay_sinh", "Ngày sinh"),
    ("quoc_tich", "Quốc tịch"),
    ("so_ho_chieu", "Số hộ chiếu"),
    ("ngay_den", "Ngày đến"),
    ("ngay_di", "Ngày đi"),
    ("dia_chi_tam_tru", "Địa chỉ tạm trú"),
    ("so_lan_nhap_canh", "Số lần NC"),
    ("tong_ngay_luu_tru_2025", "Tổng ngày (năm)"),
    ("tong_ngay_tich_luy", "Tổng ngày (tích lũy)"),
    ("trang_thai_cuoi_cung", "Mục đích/Trạng thái"),
    ("ket_qua_xac_minh", "Kết quả xác minh"),
]


def export_to_xlsx(data: List[Dict[str, Any]], filename: str = None) -> str:
    """
    Export data to XLSX file
    
    Args:
        data: List of records to export
        filename: Optional custom filename
        
    Returns:
        Path to exported file
    """
    if not data:
        raise ValueError("No data to export")
    
    # Generate filename
    if not filename:
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        filename = f"tra_cuu_{timestamp}.xlsx"
    
    # Ensure .xlsx extension
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Kết quả tra cứu"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    cell_alignment = Alignment(vertical="center", wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Status colors for conditional formatting
    status_fills = {
        "Đối tượng chú ý": PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid"),
        "Lao động": PatternFill(start_color="FFFFCC", end_color="FFFFCC", fill_type="solid"),
        "Kết hôn": PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid"),
        "Học tập": PatternFill(start_color="CCE5FF", end_color="CCE5FF", fill_type="solid"),
    }
    
    # Write headers
    for col_idx, (key, header) in enumerate(EXPORT_COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Write data
    for row_idx, record in enumerate(data, 2):
        status = record.get("trang_thai_cuoi_cung", "")
        row_fill = status_fills.get(status, None)
        
        for col_idx, (key, _) in enumerate(EXPORT_COLUMNS, 1):
            value = record.get(key, "")
            
            # Format dates
            if key in ["ngay_sinh", "ngay_den", "ngay_di"] and value:
                value = format_date_vn(value)
            
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = cell_alignment
            cell.border = thin_border
            
            # Apply status color to entire row
            if row_fill:
                cell.fill = row_fill
    
    # Adjust column widths
    column_widths = {
        1: 25,   # Họ tên
        2: 12,   # Ngày sinh
        3: 15,   # Quốc tịch
        4: 15,   # Số hộ chiếu
        5: 12,   # Ngày đến
        6: 12,   # Ngày đi
        7: 40,   # Địa chỉ
        8: 10,   # Số lần NC
        9: 12,   # Tổng ngày năm
        10: 12,  # Tổng ngày tích lũy
        11: 20,  # Mục đích
        12: 25,  # Kết quả xác minh
    }
    
    for col_idx, width in column_widths.items():
        ws.column_dimensions[chr(64 + col_idx)].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Save file
    output_path = Path(filename)
    wb.save(output_path)
    
    return str(output_path)


def export_statistics_to_xlsx(
    stats: Dict[str, Any],
    by_nationality: List[Dict],
    person_list: List[Dict],
    filters: Dict[str, Any] = None,
    filename: str = None
) -> str:
    """
    Export comprehensive statistics report to XLSX
    
    Args:
        stats: Summary statistics
        by_nationality: Statistics by nationality
        person_list: Detailed person list
        filters: Applied filters
        filename: Output filename
        
    Returns:
        Path to exported file
    """
    if not filename:
        timestamp = datetime.now().strftime("%Y-%m-%d_%H%M%S")
        filename = f"thong_ke_{timestamp}.xlsx"
    
    if not filename.endswith('.xlsx'):
        filename += '.xlsx'
    
    wb = Workbook()
    
    # ========== Sheet 1: Summary ==========
    ws1 = wb.active
    ws1.title = "Tổng hợp"
    
    # Title
    ws1.merge_cells('A1:D1')
    ws1['A1'] = "BÁO CÁO THỐNG KÊ NGƯỜI NƯỚC NGOÀI"
    ws1['A1'].font = Font(bold=True, size=14)
    ws1['A1'].alignment = Alignment(horizontal="center")
    
    # Filters info
    row = 3
    if filters:
        ws1.cell(row=row, column=1, value="Điều kiện lọc:")
        ws1['A3'].font = Font(bold=True)
        row += 1
        for key, value in filters.items():
            if value:
                ws1.cell(row=row, column=1, value=f"  - {key}: {value}")
                row += 1
        row += 1
    
    # Summary stats
    ws1.cell(row=row, column=1, value="Thống kê tổng hợp")
    ws1.cell(row=row, column=1).font = Font(bold=True)
    row += 1
    
    summary_items = [
        ("Tổng số người", stats.get("total_persons", 0)),
        ("Số quốc tịch", stats.get("total_nationalities", 0)),
        ("Đang lưu trú", stats.get("currently_residing", 0)),
        ("Lao động", stats.get("labor_count", 0)),
        ("Kết hôn", stats.get("marriage_count", 0)),
        ("Học tập", stats.get("student_count", 0)),
        ("Đối tượng chú ý", stats.get("watchlist_count", 0)),
        ("TB ngày lưu trú", stats.get("avg_days", 0)),
    ]
    
    for label, value in summary_items:
        ws1.cell(row=row, column=1, value=label)
        ws1.cell(row=row, column=2, value=value)
        row += 1
    
    # ========== Sheet 2: By Nationality ==========
    ws2 = wb.create_sheet("Theo quốc tịch")
    
    headers = ["STT", "Quốc tịch", "Số lượng", "Đang lưu trú"]
    for col, header in enumerate(headers, 1):
        ws2.cell(row=1, column=col, value=header)
        ws2.cell(row=1, column=col).font = Font(bold=True)
    
    for idx, nat in enumerate(by_nationality, 1):
        ws2.cell(row=idx+1, column=1, value=idx)
        ws2.cell(row=idx+1, column=2, value=nat.get("quoc_tich", ""))
        ws2.cell(row=idx+1, column=3, value=nat.get("count", 0))
        ws2.cell(row=idx+1, column=4, value=nat.get("still_here", 0))
    
    # ========== Sheet 3: Person List ==========
    ws3 = wb.create_sheet("Danh sách chi tiết")
    
    # Use same format as export_to_xlsx
    for col_idx, (key, header) in enumerate(EXPORT_COLUMNS, 1):
        ws3.cell(row=1, column=col_idx, value=header)
        ws3.cell(row=1, column=col_idx).font = Font(bold=True)
    
    for row_idx, record in enumerate(person_list, 2):
        for col_idx, (key, _) in enumerate(EXPORT_COLUMNS, 1):
            value = record.get(key, "")
            if key in ["ngay_sinh", "ngay_den", "ngay_di"] and value:
                value = format_date_vn(value)
            ws3.cell(row=row_idx, column=col_idx, value=value)
    
    # Adjust column widths for all sheets
    for ws in [ws1, ws2, ws3]:
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width
    
    # Save
    output_path = Path(filename)
    wb.save(output_path)
    
    return str(output_path)


def generate_template() -> str:
    """
    Generate import template Excel file
    
    Returns:
        Path to template file
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Template"
    
    headers = [
        "so_ho_chieu",
        "ho_ten",
        "ngay_sinh",
        "quoc_tich",
        "ngay_den",
        "ngay_di",
        "dia_chi_tam_tru",
        "ket_qua_xac_minh"
    ]
    
    header_vn = [
        "Số hộ chiếu",
        "Họ và tên",
        "Ngày sinh (DD/MM/YYYY)",
        "Quốc tịch",
        "Ngày đến (DD/MM/YYYY)",
        "Ngày đi (DD/MM/YYYY)",
        "Địa chỉ tạm trú",
        "Kết quả xác minh"
    ]
    
    for col, (key, vn) in enumerate(zip(headers, header_vn), 1):
        ws.cell(row=1, column=col, value=vn)
        ws.cell(row=1, column=col).font = Font(bold=True)
    
    # Sample data
    sample = [
        "E1234567", "NGUYEN VAN A", "01/01/1990", "CHINA",
        "15/01/2025", "", "123 Đường ABC, Quận XYZ", ""
    ]
    
    for col, value in enumerate(sample, 1):
        ws.cell(row=2, column=col, value=value)
    
    filename = "mau_nhap_lieu.xlsx"
    wb.save(filename)
    
    return filename
